import argparse
import csv
import os
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


START_ROW_IDX_1BASED = 4  # Excel line 4
END_ROW_IDX_1BASED = 100  # Excel line 100 (inclusive)
COL_LASTNAME_LETTER = "C"  # Column C
COL_FIRSTNAME_LETTER = "D"  # Column D
COL_SCORE_LETTER = "I"  # Column I
COL_POINTS_LETTER = "K"  # Column K
TOP_K = 15


def cell(ws, row: int, col_letter: str):
    return ws[f"{col_letter}{row}"].value


def parse_excel_all_sheets(excel_path: str) -> Dict[Tuple[str, str], Tuple[List[float], List[float]]]:
    """Parse all sheets with identical structure and aggregate scores and points by (last, first).

    Only rows 4..100 (inclusive) are considered per sheet. Columns are positional:
    - C: last name
    - D: first name
    - I: score
    - K: points

    Returns mapping player_key -> (list of scores, list of points) accumulated across sheets.
    Players without any score are not included.
    """
    wb = load_workbook(excel_path, data_only=True)

    player_to_data: Dict[Tuple[str, str], Tuple[List[float], List[float]]] = {}

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        start = START_ROW_IDX_1BASED
        end = min(END_ROW_IDX_1BASED, max_row)
        if start > end:
            continue

        for r in range(start, end + 1):
            ln = cell(ws, r, COL_LASTNAME_LETTER)
            fn = cell(ws, r, COL_FIRSTNAME_LETTER)
            sc = cell(ws, r, COL_SCORE_LETTER)
            pt = cell(ws, r, COL_POINTS_LETTER)

            if (ln is None or str(ln).strip() == "") and (fn is None or str(fn).strip() == ""):
                continue
            if sc is None:
                continue

            try:
                numeric_score = float(sc)
            except Exception:
                continue

            # Points can be None/missing, treat as 0
            try:
                numeric_points = float(pt) if pt is not None else 0.0
            except Exception:
                numeric_points = 0.0

            last_name = str(ln).strip() if ln is not None else ""
            first_name = str(fn).strip() if fn is not None else ""
            if not last_name and not first_name:
                continue

            key = (last_name, first_name)
            if key not in player_to_data:
                player_to_data[key] = ([], [])
            player_to_data[key][0].append(numeric_score)
            player_to_data[key][1].append(numeric_points)

    # Keep only players with at least one score
    return {k: v for k, v in player_to_data.items() if v[0]}


def compute_top_k_and_totals(player_data: Dict[Tuple[str, str], Tuple[List[float], List[float]]], k: int):
    # Collect with metadata for ranking
    rows_with_meta: List[Tuple[List, int, float, float, str, str]] = []  # (base_row, play_count, total_score, total_points, ln, fn)

    for (last_name, first_name), (scores, points) in player_data.items():
        # Sort by points desc to get top k points and corresponding scores
        combined = list(zip(points, scores))
        combined.sort(key=lambda x: (-x[0], -x[1]))  # points desc, then score desc
        
        top_combined = combined[:k]
        top_points = [p for p, s in top_combined]
        top_scores = [s for p, s in top_combined]
        
        total_score = sum(top_scores)
        total_points = sum(top_points)
        play_count = len(top_combined)

        # base_row in desired order except rank: [ln, fn, play_count, k points padded, total_score, total_points]
        base_row = [last_name, first_name, play_count] + top_points
        while len(base_row) < 3 + k:
            base_row.append("")
        base_row.extend([total_score, total_points])
        rows_with_meta.append((base_row, play_count, total_score, total_points, last_name, first_name))

    # Headers in final desired order
    points_cols = [str(i + 1) for i in range(k)]
    headers = ["Classement", "Nom", "Prénom", "Participations"] + points_cols + ["Totaux"] + ["Scores", "Points"]

    # Sort for ranking: by total_points desc, then total_score desc, then name for stability
    rows_with_meta.sort(key=lambda t: (-t[3], -t[2], t[4], t[5]))

    # Assign ranks with ties on (total_points, total_score)
    classement = 0
    last_key = None
    ranked_rows: List[List] = []
    for i, (base_row, play_count, total_score, total_points, ln, fn) in enumerate(rows_with_meta):
        key = (total_points, total_score)
        if last_key is None or key != last_key:
            classement = i + 1
            last_key = key
        # Final row: [rank, ln, fn, play_count, scores..., total_score, total_points]
        final_row = [classement] + base_row
        ranked_rows.append(final_row)

    # Final sort by rank asc, then name
    ranked_rows.sort(key=lambda r: (r[0], r[1], r[2]))

    return headers, ranked_rows


def export_csv(headers: List[str], rows: List[List], out_dir: str, filename: str = "classement_tarot.csv") -> str:
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, filename)
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in rows:
            padded = row[:]
            while len(padded) < len(headers):
                padded.append("")
            writer.writerow(padded)
    return out_path


def export_pdf(headers: List[str], rows: List[List], out_dir: str, filename: str = "classement_tarot.pdf", day: str = "Mardi") -> str:
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, filename)

    styles = getSampleStyleSheet()
    title = Paragraph(f"Challenge du {day}", styles["Title"])
    spacer = Spacer(1, 0.3*cm)

    # Two header rows: group label over points columns only
    k = TOP_K
    # Points start at column index 4 (0:Classement,1:Nom,2:Prénom,3:Participations)
    first_header = ["Classement", "Nom", "Prénom", "Participations", "Points"] + [""] * (k - 1) + ["Totaux"]
    second_header = ["", "", "", ""] + [str(i + 1) for i in range(k)] + ["Scores", "Points"]

    data: List[List[str]] = [first_header, second_header]

    for row in rows:
        formatted = []
        for val in row:
            if isinstance(val, (int, float)):
                if float(val).is_integer():
                    formatted.append(str(int(val)))
                else:
                    formatted.append(f"{float(val):.1f}")
            else:
                formatted.append(str(val))
        data.append(formatted)

    doc = SimpleDocTemplate(
        out_path,
        pagesize=landscape(A4),
        leftMargin=0.7*cm,
        rightMargin=0.7*cm,
        topMargin=0.7*cm,
        bottomMargin=0.7*cm,
        title="Classement Tarot",
        author="Classement Tarot",
    )

    # Column widths tuned to fit
    col_widths = []
    for h in headers:
        if h == "Classement":
            col_widths.append(1.6 * cm)
        elif h in ("Nom", "Prénom"):
            col_widths.append(2.6 * cm)
        elif h == "Participations":
            col_widths.append(1.8 * cm)
        elif h == "Totaux":
            col_widths.append(1.6 * cm)
        else:
            col_widths.append(1.2 * cm)

    table = Table(data, colWidths=col_widths, repeatRows=2)
    table.setStyle(
        TableStyle([
            ("SPAN", (4, 0), (4 + k - 1, 0)),  # Group "Points" across k columns
            ("SPAN", (4 + k, 0), (4 + k + 1, 0)),  # Group "Totaux" across Scores and Points columns
            ("BACKGROUND", (0, 0), (-1, 1), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 1), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 2), (2, -1), "LEFT"),  # left-align Nom & Prénom in body
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.whitesmoke, colors.lightcyan]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )

    story = [title, spacer, table]
    doc.build(story)
    return out_path


def run(excel_path: str, out_dir: str, want_pdf: bool, want_csv: bool, day: str):
    player_data = parse_excel_all_sheets(excel_path)
    headers, rows = compute_top_k_and_totals(player_data, TOP_K)

    outputs: Dict[str, str] = {}
    if want_csv:
        outputs["csv"] = export_csv(headers, rows, out_dir, f"classement_tarot_{day}.csv")
    if want_pdf:
        outputs["pdf"] = export_pdf(headers, rows, out_dir, f"classement_tarot_{day}.pdf", day)
    return outputs


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Agrège un Excel multi-feuilles (tournois) en classement des joueurs, "
            "avec top 15, total et classement, puis export CSV/PDF paysage."
        )
    )
    parser.add_argument("excel", help="Chemin vers le fichier Excel (.xlsx)")
    parser.add_argument("--out", dest="out", default=".", help="Dossier de sortie (défaut: .)")
    parser.add_argument("--pdf", dest="pdf", action="store_true", help="Générer un PDF paysage")
    parser.add_argument("--csv", dest="csv", action="store_true", help="Générer un CSV")
    parser.add_argument("--day", dest="day", default="Mardi", help="Jour du tournoi")
    return parser


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()

    out_dir = args.out
    want_pdf = bool(args.pdf)
    want_csv = bool(args.csv)
    day = args.day

    if not want_pdf and not want_csv:
        want_pdf = True
        want_csv = True

    outputs = run(args.excel, out_dir, want_pdf, want_csv, day)

    for kind, path in outputs.items():
        print(f"Export {kind.upper()}: {path}")


if __name__ == "__main__":
    main()
